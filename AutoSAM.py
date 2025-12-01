import requests
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re

# ==== FUNCTION: LOAD CONFIG FROM EXCEL ====
def load_config_from_excel(path="config.xlsx"):
    wb = load_workbook(path)

    # Load Allowed Software
    ws_allowed = wb["Allowed"]
    allowed = [row[0].strip() for row in ws_allowed.iter_rows(min_row=2, values_only=True) if row[0]]

    # Load Excluded Keywords
    ws_excluded = wb["Excluded"]
    excluded = [row[0].strip() for row in ws_excluded.iter_rows(min_row=2, values_only=True) if row[0]]

    # Load User–Department Mapping
    ws_map = wb["UserDeptMap"]
    mapping = {}
    for row in ws_map.iter_rows(min_row=2, values_only=True):
        system, user, dept = row
        if system:
            mapping[system.strip()] = (user.strip() if user else "N/A", dept.strip() if dept else "N/A")

    return allowed, excluded, mapping

# ==== CONFIGURATION PLACEHOLDERS ====
API_URL = "https://YOUR_GLPI_INSTANCE/api"
APP_TOKEN = "YOUR_APP_TOKEN"
USER_TOKEN = "YOUR_USER_TOKEN"
OUTPUT_FILE = r"path\to\SAM_report.xlsx"
CONFIG_FILE = r"path\to\config.xlsx"

# Load configuration from Excel
ALLOWED_SOFTWARE, EXCLUDED_KEYWORDS, USER_DEPT_MAP = load_config_from_excel(CONFIG_FILE)

# ==== FUNCTION: EXCLUDE SOFTWARE ====
def is_excluded(name: str) -> bool:
    lname = name.lower().strip()
    # Ignore Windows Updates like KB5066613, KB5054156, etc.
    if re.search(r'\bkb\d{5,7}\b', lname):
        return True
    return any(keyword.lower() in lname for keyword in EXCLUDED_KEYWORDS)

# ==== FUNCTION: CLEAN SOFTWARE NAME ====
def clean_name(name: str) -> str:
    name = name.strip()
    if not name or name.startswith("{") or len(name) > 40:
        return None
    return name

# ==== FUNCTION: CHECK IF SOFTWARE IS ALLOWED (INCLUDING VISUAL STUDIO) ====
def is_allowed(sw: str) -> bool:
    sw_lower = sw.lower()
    if "vs" in sw_lower or "visual studio" in sw_lower:
        return True
    for allowed in ALLOWED_SOFTWARE:
        if allowed.lower() in sw_lower or sw_lower in allowed.lower():
            return True
    return False

# ==== FUNCTION: FETCH PAGINATED GLPI DATA ====
def fetch_paginated(url, headers, step=1000):
    results = []
    start = 0
    while True:
        paged_url = f"{url}&range={start}-{start+step-1}"
        r = requests.get(paged_url, headers=headers)
        r.raise_for_status()
        data = r.json()
        if not data:
            break
        chunk = data["data"] if isinstance(data, dict) and "data" in data else data
        results.extend(chunk)
        if len(chunk) < step:
            break
        start += step
    return results

# ==== START GLPI SESSION ====
r = requests.post(
    f"{API_URL}/initSession",
    headers={"App-Token": APP_TOKEN, "Content-Type": "application/json"},
    json={"user_token": USER_TOKEN}
)
r.raise_for_status()
session = r.json()["session_token"]
headers = {"App-Token": APP_TOKEN, "Session-Token": session}
print("✅ Session started with GLPI")

# ==== FETCH COMPUTERS ====
computers = fetch_paginated(f"{API_URL}/Computer?is_deleted=0", headers)
print(f"💻 Retrieved {len(computers)} computers from GLPI")

# ==== PREPARE EXCEL ====
wb = openpyxl.Workbook()
ws_main = wb.active
ws_main.title = "Software Audit"
ws_unauth = wb.create_sheet("Unauthorized Software")

# Headers
headers_row = ["S. No.", "System Name", "User Name", "Department"] + ALLOWED_SOFTWARE
unauth_headers = ["S. No.", "System Name", "User Name", "Department", "Unauthorized Software"]

ws_main.append(headers_row)
ws_unauth.append(unauth_headers)

# Styling setup
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))

for ws in [ws_main, ws_unauth]:
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

# ==== PROCESS EACH COMPUTER ====
checkmark = "✓"
unauth_index = 1

for idx, comp in enumerate(computers, start=1):
    comp_id = comp["id"]
    comp_name = comp.get("name", f"Computer-{comp_id}")
    print(f"🔍 Processing {idx}/{len(computers)}: {comp_name}")

    # Fetch installed software
    search_url = (
        f"{API_URL}/search/Software"
        f"?is_deleted=0&as_map=0"
        f"&criteria[0][link]=AND"
        f"&criteria[0][itemtype]=Computer"
        f"&criteria[0][meta]=1"
        f"&criteria[0][field]=2"
        f"&criteria[0][searchtype]=contains"
        f"&criteria[0][value]=^{comp_id}$"
    )
    software_entries = fetch_paginated(search_url, headers)
    installed = []

    # Extract and clean names
    for rel in software_entries:
        sw_name = rel.get("1", "").strip()
        if not sw_name:
            continue
        cleaned = clean_name(sw_name)
        if cleaned and not is_excluded(cleaned):
            installed.append(cleaned.lower())

    # Main sheet row
    user, dept = USER_DEPT_MAP.get(comp_name, ("N/A", "N/A"))
    row_data = [idx, comp_name, user, dept]
    unauthorized_list = []

    # Check allowed software (including VS handling)
    for allowed in ALLOWED_SOFTWARE:
        if allowed.lower() == "visual studio":
            match_found = any("vs" in sw or "visual studio" in sw for sw in installed)
        else:
            match_found = any(allowed.lower() in sw or sw in allowed.lower() for sw in installed)
        row_data.append(checkmark if match_found else "")

    # Find unauthorized software
    for sw in installed:
        if not is_allowed(sw):
            unauthorized_list.append(sw)

    ws_main.append(row_data)

    # Fill unauthorized sheet
    for sw in unauthorized_list:
        unauth_index += 1
        ws_unauth.append([unauth_index - 1, comp_name, user, dept, sw])

# ==== STYLING & FORMATTING ====
for ws in [ws_main, ws_unauth]:
    for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
    # Auto column width
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2
    ws.freeze_panes = "A2"

# ==== SAVE EXCEL ====
wb.save(OUTPUT_FILE)
print(f"✅ Report generated successfully at: {OUTPUT_FILE}")

# ==== END GLPI SESSION ====
requests.post(f"{API_URL}/killSession", headers=headers)
print("🔒 GLPI session closed.")
